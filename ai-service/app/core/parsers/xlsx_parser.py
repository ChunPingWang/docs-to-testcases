from openpyxl import load_workbook
from pathlib import Path


def parse_xlsx(file_path: str) -> dict:
    """Parse an .xlsx file and extract text from all sheets."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        header = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if not any(cells):
                continue
            if i == 0:
                header = cells
            rows.append(cells)

        if not rows:
            continue

        # Format as readable text
        if header:
            text_lines = [" | ".join(header)]
            text_lines.append("-" * len(text_lines[0]))
            for row in rows[1:]:
                text_lines.append(" | ".join(row))
        else:
            text_lines = [" | ".join(row) for row in rows]

        sections.append({
            "title": sheet_name,
            "heading_path": sheet_name,
            "content": "\n".join(text_lines),
        })

    wb.close()

    full_text = "\n\n".join(
        f"## {s['title']}\n{s['content']}" for s in sections
    )

    return {
        "text": full_text,
        "sections": sections,
        "tables": [],
        "metadata": {
            "filename": Path(file_path).name,
            "sheet_count": len(sections),
        },
    }
